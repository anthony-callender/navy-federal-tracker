"""
Excel Manager - Handles reading/writing transactions to Excel file
"""

from pathlib import Path
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


class ExcelManager:
    """Manages the transactions Excel file."""
    
    # Column configuration
    COLUMNS = [
        ("Date", 12),
        ("Time", 10),
        ("Description", 40),
        ("Merchant", 30),
        ("Amount", 12),
        ("Type", 10),
        ("Category", 18),
        ("Account", 20),
        ("Account Number", 15),
        ("Email ID", 15),
    ]
    
    # Styling
    HEADER_FILL = PatternFill('solid', fgColor='4472C4')
    CREDIT_FILL = PatternFill('solid', fgColor='C6EFCE')  # Light green
    DEBIT_FILL = PatternFill('solid', fgColor='FFC7CE')   # Light red
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, filepath: Optional[Path] = None):
        self.filepath = filepath or config.TRANSACTIONS_FILE
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Create the Excel file with headers if it doesn't exist."""
        # Ensure directory exists
        self.filepath.parent.mkdir(parents=True, exist_ok=True)
        
        if not self.filepath.exists():
            self._create_new_file()
    
    def _create_new_file(self):
        """Create a new Excel file with headers and formatting."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Add headers
        for col, (header, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add a Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "Transaction Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        wb.save(self.filepath)
        print(f"✅ Created new transactions file: {self.filepath}")
    
    def add_transactions(self, transactions: List) -> int:
        """
        Add transactions to the Excel file.
        
        Args:
            transactions: List of Transaction objects
            
        Returns:
            Number of new transactions added
        """
        if not transactions:
            return 0
        
        wb = load_workbook(self.filepath)
        ws = wb["Transactions"]
        
        # Get existing email IDs to avoid duplicates
        existing_ids = self._get_existing_email_ids(ws)
        
        added = 0
        for tx in transactions:
            # Skip if already exists
            if tx.email_id in existing_ids:
                continue
            
            # Add to next row
            row = ws.max_row + 1
            tx_dict = tx.to_dict()
            
            for col, (header, _) in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=row, column=col, value=tx_dict.get(header, ""))
                cell.border = self.THIN_BORDER
                
                # Format amount column
                if header == "Amount":
                    cell.number_format = '$#,##0.00'
                    if tx_dict.get("Amount", 0) >= 0:
                        cell.fill = self.CREDIT_FILL
                    else:
                        cell.fill = self.DEBIT_FILL
            
            added += 1
            existing_ids.add(tx.email_id)
        
        if added > 0:
            wb.save(self.filepath)
            self._update_summary(wb)
            print(f"✅ Added {added} new transactions")
        
        return added
    
    def _get_existing_email_ids(self, ws) -> set:
        """Get set of email IDs already in the spreadsheet."""
        email_id_col = None
        for col, (header, _) in enumerate(self.COLUMNS, 1):
            if header == "Email ID":
                email_id_col = col
                break
        
        if not email_id_col:
            return set()
        
        ids = set()
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=email_id_col).value
            if cell_value:
                ids.add(str(cell_value))
        
        return ids
    
    def _update_summary(self, wb: Workbook):
        """Update the summary sheet with totals."""
        ws_tx = wb["Transactions"]
        ws_summary = wb["Summary"]
        
        # Clear existing summary data
        for row in range(3, 20):
            for col in range(1, 5):
                ws_summary.cell(row=row, column=col).value = None
        
        # Calculate totals by category
        category_totals = {}
        total_income = 0
        total_expenses = 0
        
        amount_col = None
        category_col = None
        for col, (header, _) in enumerate(self.COLUMNS, 1):
            if header == "Amount":
                amount_col = col
            elif header == "Category":
                category_col = col
        
        for row in range(2, ws_tx.max_row + 1):
            amount = ws_tx.cell(row=row, column=amount_col).value or 0
            category = ws_tx.cell(row=row, column=category_col).value or "Other"
            
            if amount > 0:
                total_income += amount
            else:
                total_expenses += abs(amount)
            
            category_totals[category] = category_totals.get(category, 0) + amount
        
        # Write summary
        ws_summary['A3'] = "Total Income:"
        ws_summary['B3'] = total_income
        ws_summary['B3'].number_format = '$#,##0.00'
        ws_summary['B3'].fill = self.CREDIT_FILL
        
        ws_summary['A4'] = "Total Expenses:"
        ws_summary['B4'] = total_expenses
        ws_summary['B4'].number_format = '$#,##0.00'
        ws_summary['B4'].fill = self.DEBIT_FILL
        
        ws_summary['A5'] = "Net:"
        ws_summary['B5'] = total_income - total_expenses
        ws_summary['B5'].number_format = '$#,##0.00'
        ws_summary['B5'].font = Font(bold=True)
        
        ws_summary['A7'] = "By Category:"
        ws_summary['A7'].font = Font(bold=True)
        
        row = 8
        for category, total in sorted(category_totals.items(), key=lambda x: x[1]):
            ws_summary.cell(row=row, column=1).value = category
            ws_summary.cell(row=row, column=2).value = total
            ws_summary.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        ws_summary['A1'] = f"Transaction Summary (Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')})"
        
        wb.save(self.filepath)
    
    def get_last_sync_date(self) -> Optional[datetime]:
        """Get the date of the most recent transaction."""
        try:
            wb = load_workbook(self.filepath)
            ws = wb["Transactions"]
            
            if ws.max_row < 2:
                return None
            
            # Find date column
            date_col = 1  # Assuming first column is Date
            
            latest = None
            for row in range(2, ws.max_row + 1):
                date_str = ws.cell(row=row, column=date_col).value
                if date_str:
                    try:
                        if isinstance(date_str, datetime):
                            date = date_str
                        else:
                            date = datetime.strptime(str(date_str), "%Y-%m-%d")
                        if latest is None or date > latest:
                            latest = date
                    except:
                        pass
            
            return latest
        except:
            return None
    
    def get_transaction_count(self) -> int:
        """Get total number of transactions."""
        try:
            wb = load_workbook(self.filepath)
            ws = wb["Transactions"]
            return max(0, ws.max_row - 1)  # Subtract header row
        except:
            return 0


# Quick test
if __name__ == "__main__":
    manager = ExcelManager()
    print(f"Transactions file: {manager.filepath}")
    print(f"Transaction count: {manager.get_transaction_count()}")
    print(f"Last sync: {manager.get_last_sync_date()}")
